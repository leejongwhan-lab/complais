"""Load `esg_master_kpis` from ESG_KPI_최종통합본_353개.xlsx (idempotent reload)."""
from __future__ import annotations

import re
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import load_workbook
from sqlalchemy import text
from sqlalchemy.orm import Session

from app.models.esg import EsgMasterKpi

BASE_DIR = Path(__file__).resolve().parents[2]

# NFD/NFC-safe discovery under backend root
def _default_excel() -> Path:
    for p in sorted(BASE_DIR.glob("ESG_KPI*.xlsx")):
        if "353" in p.name or "최종" in p.name or "최종" in p.name:
            return p
    matches = list(BASE_DIR.glob("ESG_KPI*.xlsx"))
    if matches:
        return matches[0]
    return BASE_DIR / "ESG_KPI_최종통합본_353개.xlsx"


DEFAULT_EXCEL = _default_excel()

ESG_CATEGORY_MAP = {
    "환경": "E",
    "사회": "S",
    "거버넌스": "G",
    "지배구조": "G",
}

# Excel "공통" rows: infer E/S/G from KPI name / ISO label (ENUM has no Common)
_COMMON_E = ("환경", "에너지", "온실", "배출", "폐기물", "용수", "수자원")
_COMMON_S = ("안전", "보건", "노동", "고용", "인권", "다양성")


def _infer_common_category(kpi_name: str, standard: str) -> str:
    text = f"{kpi_name} {standard}"
    if any(k in text for k in _COMMON_E) or "14001" in text or "50001" in text:
        return "E"
    if any(k in text for k in _COMMON_S) or "45001" in text:
        return "S"
    return "G"


def _map_esg_category(raw: str, kpi_name: str, standard: str) -> str:
    raw = (raw or "").strip()
    if raw in ESG_CATEGORY_MAP:
        return ESG_CATEGORY_MAP[raw]
    if raw == "공통":
        return _infer_common_category(kpi_name, standard)
    raise ValueError(f"Unknown ESG영역: {raw!r}")


def _source_codes(raw: str) -> Tuple[str, str]:
    """Return (source_type_code, extraction_detail_method).

    Excel has A/B/C/D 출처 분류 (not A-1~C-2 detail). Store letter + full label.
    """
    text = (raw or "").strip() or "D 기업작성형"
    m = re.match(r"^([A-Da-d])\b", text)
    code = m.group(1).upper() if m else text[:20]
    return code, text[:100]


def _as_bool(raw: str, true_tokens: Tuple[str, ...]) -> bool:
    text = (raw or "").strip()
    return any(t in text for t in true_tokens)


def parse_esg_kpi_excel(excel_path: Path | str = DEFAULT_EXCEL) -> List[Dict[str, Any]]:
    excel_path = Path(excel_path)
    if not excel_path.exists():
        raise FileNotFoundError(f"ESG KPI Excel not found: {excel_path}")

    wb = load_workbook(excel_path, read_only=True, data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    header = next(rows_iter)
    header_map = {str(h).strip(): i for i, h in enumerate(header) if h is not None}

    def col(*names: str) -> Optional[int]:
        for n in names:
            if n in header_map:
                return header_map[n]
        return None

    idx = {
        "esg": col("ESG영역"),
        "sub": col("세부카테고리"),
        "name": col("KPI/항목명"),
        "quant": col("정량/정성"),
        "unit": col("단위/형식"),
        "source": col("데이터 출처 분류"),
        "criteria": col("ISO/기준 매핑"),
        "desc": col("정의/수집설명"),
        "audit": col("ISO 심사 검증 여부"),
        "std": col("ISO 표준"),
        "clause": col("ISO 조항"),
        "api": col("공공데이터 자동수집 여부"),
    }
    missing = [k for k, v in idx.items() if v is None and k != "criteria"]
    if missing:
        wb.close()
        raise ValueError(f"Missing Excel columns: {missing}; got {list(header_map)}")

    out: List[Dict[str, Any]] = []
    for row in rows_iter:
        def cell(key: str) -> str:
            col_i = idx.get(key)
            if col_i is None:
                return ""
            v = row[col_i]
            if v is None:
                return ""
            return str(v).strip()

        kpi_name = cell("name")
        if not kpi_name:
            continue
        standard = cell("std") or "기타/확인필요"
        clause = cell("clause") or "-"
        unit = cell("unit") or "-"
        source_code, extraction = _source_codes(cell("source"))
        quant_raw = cell("quant")
        out.append(
            {
                "esg_category": _map_esg_category(cell("esg"), kpi_name, standard),
                "sub_category": cell("sub") or "미분류",
                "kpi_name": kpi_name[:200],
                "is_quantitative": quant_raw == "정량" or "정량" in quant_raw,
                "unit_format": unit[:50],
                "managed_standard_name": standard[:150],
                "iso_clause_detail": clause[:150],
                "is_iso_auditable": _as_bool(
                    cell("audit"), ("검증 가능", "심사 검증")
                ),
                "source_type_code": source_code[:20],
                "extraction_detail_method": extraction,
                "is_public_api_available": _as_bool(
                    cell("api"), ("자동수집 가능",)
                ),
                "criteria_mapping": (cell("criteria") or None),
                "description": cell("desc") or None,
            }
        )
        if out[-1]["criteria_mapping"]:
            out[-1]["criteria_mapping"] = out[-1]["criteria_mapping"][:150]
    wb.close()
    return out


def reload_esg_master_kpis(
    db: Session,
    excel_path: Path | str = DEFAULT_EXCEL,
) -> Dict[str, int]:
    """Clear table and reload from Excel. Idempotent full replace."""
    rows = parse_esg_kpi_excel(excel_path)
    deleted = db.query(EsgMasterKpi).delete()
    db.flush()
    # InnoDB keeps AUTO_INCREMENT after DELETE — reset so KPI-E-001 style ids return.
    db.execute(text("ALTER TABLE esg_master_kpis AUTO_INCREMENT = 1"))
    for raw in rows:
        db.add(EsgMasterKpi(**raw))
    db.commit()
    return {"deleted": int(deleted or 0), "inserted": len(rows)}
