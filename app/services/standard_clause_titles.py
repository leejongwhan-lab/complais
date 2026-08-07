"""Authoritative per-standard clause_no + official title.

Seed source: 스토리보드 sheet 「표준별조항번호 및 제목」
Center of truth: ``standard_clause_masters`` (FK → ``standard_masters``).

Do not invent titles. Skip PENDING editions (empty Excel columns).
Never DROP companies / CB masters — only ADD/UPDATE clause rows.
"""
from __future__ import annotations

import logging
import os
import re
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import load_workbook
from sqlalchemy import text
from sqlalchemy.orm import Session

from app.data.standards_catalog import STANDARD_CATALOG

logger = logging.getLogger(__name__)

BASE_DIR = Path(__file__).resolve().parents[2]
SHEET_NAME_ALIASES = (
    "표준별조항번호 및 제목",
    "표준별 조항번호 및 제목",
    "표준 별 조항번호 및 조항제목",
)


def _candidate_excels() -> List[Path]:
    out: List[Path] = []
    env = os.environ.get("CLAUSE_TITLE_EXCEL")
    if env:
        out.append(Path(env))
    out.extend(
        [
            BASE_DIR / "스토리보드_0721.xlsx",
            BASE_DIR / "data" / "스토리보드_0721.xlsx",
            BASE_DIR / "표준 별 조항번호 및 조항제목.xlsx",
        ]
    )
    return out


def resolve_clause_title_excel() -> Optional[Path]:
    for p in _candidate_excels():
        if p.exists():
            return p
    return None


def _table_exists(db: Session, name: str) -> bool:
    row = db.execute(
        text(
            "SELECT 1 FROM information_schema.tables "
            "WHERE table_schema = DATABASE() AND table_name = :t LIMIT 1"
        ),
        {"t": name},
    ).first()
    return bool(row)


def _cell(v: Any) -> Optional[str]:
    if v is None:
        return None
    s = str(v).strip()
    return s if s else None


def _normalize_clause_no(raw: str) -> str:
    s = str(raw).strip()
    # openpyxl may yield 4.1 as float → "4.1"; also "4.0" → "4"
    if re.fullmatch(r"\d+\.0+", s):
        return s.split(".", 1)[0]
    return s


def _clause_sort_key(clause_no: str) -> Tuple:
    parts = []
    for p in re.split(r"[.\-/~]", str(clause_no)):
        if p.isdigit():
            parts.append((0, int(p)))
        else:
            parts.append((1, p))
    return tuple(parts) if parts else ((1, str(clause_no)),)


def _match_catalog(code: str) -> Optional[Any]:
    compact = code.replace(" ", "").upper()
    # exact display match first
    for s in STANDARD_CATALOG:
        if s.display_code.replace(" ", "").upper() == compact:
            return s
    # ISO 9001:2015 style
    year_m = re.search(r":(\d{4})", code)
    year = int(year_m.group(1)) if year_m else None
    iso_m = re.search(r"(ISO(?:/IEC)?\s*\d{4,5})", code, re.IGNORECASE)
    iso_part = re.sub(r"\s+", " ", iso_m.group(1)).upper() if iso_m else None
    if iso_part and year is not None:
        for s in STANDARD_CATALOG:
            if (
                s.edition_year == year
                and s.iso_number.replace(" ", "").upper()
                == iso_part.replace(" ", "").upper()
            ):
                return s
    return None


def _parse_header(raw: str) -> Optional[Dict[str, Any]]:
    """'ISO 9001:2015 품질경영시스템' → catalog fields."""
    text_s = (raw or "").strip()
    if not text_s:
        return None
    m = re.match(r"^(ISO(?:/IEC)?\s+[\d\-:]+)\s+(.+)$", text_s)
    if m:
        code, name = m.group(1).strip(), m.group(2).strip()
    else:
        code, name = text_s, text_s

    catalog = _match_catalog(code)
    year_match = re.search(r":(\d{4})", code)
    year = (
        catalog.edition_year
        if catalog and catalog.edition_year is not None
        else (int(year_match.group(1)) if year_match else None)
    )
    if catalog is None:
        digits = re.search(r"(\d{4,5})", code)
        fam = f"ISO{digits.group(1)}" if digits else "UNK"
        return {
            "standard_key": f"{fam}_{year}" if year else fam,
            "family_code": fam,
            "edition_year": year,
            "iso_number": re.sub(r":\d{4}$", "", code).strip(),
            "display_code": code,
            "standard_name": name,
            "clauses_status": "READY",
            "clauses_note": None,
            "role": "CERTIFIABLE",
            "skip_clauses": False,
        }

    return {
        "standard_key": catalog.standard_key,
        "family_code": catalog.family_code,
        "edition_year": year,
        "iso_number": catalog.iso_number,
        "display_code": catalog.display_code,
        "standard_name": catalog.name_ko,
        "clauses_status": catalog.clauses_status,
        "clauses_note": catalog.clauses_note,
        "role": catalog.role,
        "skip_clauses": catalog.clauses_status == "PENDING",
    }


def _find_sheet(wb) -> Optional[str]:
    for name in SHEET_NAME_ALIASES:
        if name in wb.sheetnames:
            return name
    for name in wb.sheetnames:
        if "조항번호" in name and "제목" in name:
            return name
    return None


def parse_clause_title_sheet(
    excel_path: Path,
) -> Tuple[Dict[str, List[Dict[str, Any]]], Dict[str, Dict[str, Any]]]:
    """Return (clauses_by_key, metas_by_key)."""
    wb = load_workbook(excel_path, data_only=True)
    sheet = _find_sheet(wb)
    if not sheet:
        raise ValueError(f"Clause-title sheet not found in {excel_path}")
    ws = wb[sheet]

    header_row = None
    for r in range(1, min(8, (ws.max_row or 1) + 1)):
        v = ws.cell(r, 2).value
        if v and "조항" in str(v):
            header_row = r
            break
    if header_row is None:
        header_row = 3

    standards: List[Tuple[int, Dict[str, Any]]] = []
    for c in range(3, (ws.max_column or 3) + 1):
        raw = _cell(ws.cell(header_row, c).value)
        if not raw:
            continue
        meta = _parse_header(raw)
        if meta:
            standards.append((c, meta))

    out: Dict[str, List[Dict[str, Any]]] = {}
    metas: Dict[str, Dict[str, Any]] = {}
    for col, meta in standards:
        sk = meta["standard_key"]
        metas[sk] = meta
        if meta.get("skip_clauses"):
            out[sk] = []
            continue
        rows: List[Dict[str, Any]] = []
        sort_i = 0
        for r in range(header_row + 1, (ws.max_row or header_row) + 1):
            cno_raw = ws.cell(r, 2).value
            if cno_raw is None:
                continue
            cno = _normalize_clause_no(cno_raw)
            if not cno:
                continue
            title = _cell(ws.cell(r, col).value)
            if not title or title == "-":
                continue
            sort_i += 1
            rows.append(
                {
                    "clause_number": cno,
                    "clause_title_kr": title,
                    "sort_order": sort_i,
                    "depth": len(cno.split(".")),
                }
            )
        out[sk] = rows
    return out, metas


def ensure_standard_clause_tables(db: Session) -> None:
    if not _table_exists(db, "standard_masters"):
        db.execute(
            text(
                """
CREATE TABLE IF NOT EXISTS standard_masters (
  id INT NOT NULL AUTO_INCREMENT,
  standard_key VARCHAR(40) NOT NULL,
  family_code VARCHAR(20) NOT NULL,
  edition_year INT NULL,
  iso_number VARCHAR(40) NOT NULL,
  display_code VARCHAR(60) NOT NULL,
  standard_code VARCHAR(60) NOT NULL,
  standard_name VARCHAR(100) NOT NULL,
  version_year INT NULL,
  clauses_status VARCHAR(20) NOT NULL DEFAULT 'READY',
  clauses_note VARCHAR(255) NULL,
  role VARCHAR(20) NOT NULL DEFAULT 'CERTIFIABLE',
  is_active TINYINT(1) NOT NULL DEFAULT 1,
  description VARCHAR(255) NULL,
  created_at DATETIME NULL DEFAULT CURRENT_TIMESTAMP,
  updated_at DATETIME NULL,
  PRIMARY KEY (id),
  UNIQUE KEY uq_standard_key (standard_key),
  UNIQUE KEY uq_standard_code (standard_code)
) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COLLATE=utf8mb4_unicode_ci
"""
            )
        )
    if not _table_exists(db, "standard_clause_masters"):
        db.execute(
            text(
                """
CREATE TABLE IF NOT EXISTS standard_clause_masters (
  id INT NOT NULL AUTO_INCREMENT,
  standard_id INT NOT NULL,
  clause_number VARCHAR(30) NOT NULL,
  clause_title_kr VARCHAR(255) NOT NULL DEFAULT '',
  depth INT NOT NULL DEFAULT 1,
  sort_order INT NOT NULL DEFAULT 0,
  requirements_summary TEXT NULL,
  created_at DATETIME NULL DEFAULT CURRENT_TIMESTAMP,
  updated_at DATETIME NULL,
  PRIMARY KEY (id),
  UNIQUE KEY uix_standard_clause_number (standard_id, clause_number),
  KEY idx_clause_standard (standard_id)
) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COLLATE=utf8mb4_unicode_ci
"""
            )
        )
    db.commit()


def count_clause_masters(db: Session) -> int:
    if not _table_exists(db, "standard_clause_masters"):
        return 0
    return int(db.execute(text("SELECT COUNT(*) FROM standard_clause_masters")).scalar() or 0)


def _current_counts(db: Session) -> Dict[str, int]:
    if not _table_exists(db, "standard_clause_masters"):
        return {}
    rows = db.execute(
        text(
            "SELECT sm.standard_key, COUNT(c.id) AS n "
            "FROM standard_masters sm "
            "LEFT JOIN standard_clause_masters c ON c.standard_id = sm.id "
            "GROUP BY sm.standard_key"
        )
    ).fetchall()
    return {r[0]: int(r[1]) for r in rows if int(r[1]) > 0}


def seed_standard_clause_titles(
    db: Session,
    excel_path: Optional[Path] = None,
    *,
    force: bool = False,
) -> Dict[str, int]:
    """Idempotent upsert of official clause titles from storyboard Excel."""
    path = Path(excel_path) if excel_path else resolve_clause_title_excel()
    if not path or not path.exists():
        raise FileNotFoundError(
            "Clause-title Excel not found (스토리보드_0721.xlsx / CLAUSE_TITLE_EXCEL)"
        )

    ensure_standard_clause_tables(db)
    existing = count_clause_masters(db)
    if existing > 0 and not force:
        return _current_counts(db)

    by_std, metas = parse_clause_title_sheet(path)
    counts: Dict[str, int] = {}

    for sk, clauses in by_std.items():
        meta = metas.get(sk)
        if not meta:
            continue

        row = db.execute(
            text(
                "SELECT id FROM standard_masters "
                "WHERE standard_key = :sk OR standard_code = :code LIMIT 1"
            ),
            {"sk": meta["standard_key"], "code": meta["display_code"]},
        ).first()
        if row:
            std_id = int(row[0])
            db.execute(
                text(
                    "UPDATE standard_masters SET "
                    "standard_key=:sk, family_code=:fam, edition_year=:yr, "
                    "iso_number=:iso, display_code=:disp, standard_code=:code, "
                    "standard_name=:name, version_year=:yr, "
                    "clauses_status=:st, clauses_note=:note, role=:role, is_active=1 "
                    "WHERE id=:id"
                ),
                {
                    "id": std_id,
                    "sk": meta["standard_key"],
                    "fam": meta["family_code"],
                    "yr": meta["edition_year"],
                    "iso": meta["iso_number"],
                    "disp": meta["display_code"],
                    "code": meta["display_code"],
                    "name": meta["standard_name"],
                    "st": meta["clauses_status"],
                    "note": meta.get("clauses_note"),
                    "role": meta["role"],
                },
            )
        else:
            db.execute(
                text(
                    "INSERT INTO standard_masters "
                    "(standard_key, family_code, edition_year, iso_number, display_code, "
                    " standard_code, standard_name, version_year, clauses_status, "
                    " clauses_note, role, is_active) "
                    "VALUES (:sk, :fam, :yr, :iso, :disp, :code, :name, :yr, :st, :note, :role, 1)"
                ),
                {
                    "sk": meta["standard_key"],
                    "fam": meta["family_code"],
                    "yr": meta["edition_year"],
                    "iso": meta["iso_number"],
                    "disp": meta["display_code"],
                    "code": meta["display_code"],
                    "name": meta["standard_name"],
                    "st": meta["clauses_status"],
                    "note": meta.get("clauses_note"),
                    "role": meta["role"],
                },
            )
            std_id = int(
                db.execute(
                    text("SELECT id FROM standard_masters WHERE standard_key=:sk"),
                    {"sk": meta["standard_key"]},
                ).scalar()
            )

        if meta.get("skip_clauses"):
            counts[sk] = 0
            continue

        keep_numbers = set()
        for c in clauses:
            cno = c["clause_number"]
            keep_numbers.add(cno)
            db.execute(
                text(
                    "INSERT INTO standard_clause_masters "
                    "(standard_id, clause_number, clause_title_kr, depth, sort_order) "
                    "VALUES (:sid, :cno, :title, :depth, :sort) "
                    "ON DUPLICATE KEY UPDATE "
                    "clause_title_kr=VALUES(clause_title_kr), "
                    "depth=VALUES(depth), sort_order=VALUES(sort_order), "
                    "updated_at=CURRENT_TIMESTAMP"
                ),
                {
                    "sid": std_id,
                    "cno": str(cno)[:30],
                    "title": str(c["clause_title_kr"])[:255],
                    "depth": int(c["depth"]),
                    "sort": int(c["sort_order"]),
                },
            )

        # Obsolete rows: only remove when not referenced (soft; never DROP masters)
        if keep_numbers:
            existing_rows = db.execute(
                text(
                    "SELECT id, clause_number FROM standard_clause_masters "
                    "WHERE standard_id = :sid"
                ),
                {"sid": std_id},
            ).fetchall()
            obsolete = [eid for eid, enum in existing_rows if enum not in keep_numbers]
            for eid in obsolete:
                try:
                    with db.begin_nested():
                        db.execute(
                            text("DELETE FROM standard_clause_masters WHERE id = :id"),
                            {"id": eid},
                        )
                except Exception as exc:
                    logger.warning("skip delete obsolete clause id=%s: %s", eid, exc)

        counts[sk] = len(keep_numbers)

    db.commit()
    logger.info("standard_clause_masters seeded from %s: %s", path.name, counts)
    return counts


def ensure_standard_clause_titles(
    db: Session, *, force: bool = False
) -> Dict[str, int]:
    """Ensure table populated; seed from Excel when empty or force=True."""
    ensure_standard_clause_tables(db)
    if count_clause_masters(db) == 0 or force:
        try:
            return seed_standard_clause_titles(db, force=True)
        except FileNotFoundError as exc:
            logger.warning("clause title seed skipped: %s", exc)
            return {}
    return seed_standard_clause_titles(db, force=False)


def list_official_clauses(
    db: Session, standard_key: str
) -> List[Dict[str, Any]]:
    """Official clause_no + title for a platform standard_key (e.g. QMS_2015)."""
    if not _table_exists(db, "standard_clause_masters"):
        return []

    from app.services.iso_clauses_master import resolve_standard_key

    sk = resolve_standard_key(standard_key) or standard_key
    row = db.execute(
        text(
            "SELECT id, standard_key, family_code, display_code FROM standard_masters "
            "WHERE standard_key = :sk OR standard_code = :raw OR display_code = :raw "
            "LIMIT 1"
        ),
        {"sk": sk, "raw": standard_key},
    ).mappings().first()
    if not row:
        return []

    rows = db.execute(
        text(
            "SELECT clause_number, clause_title_kr, depth, sort_order "
            "FROM standard_clause_masters WHERE standard_id = :sid "
            "ORDER BY sort_order, clause_number"
        ),
        {"sid": row["id"]},
    ).mappings().all()
    out = []
    for r in rows:
        out.append(
            {
                "standard_key": row["standard_key"],
                "family_code": row["family_code"],
                "display_code": row["display_code"],
                "clause_no": r["clause_number"],
                "clause_number": r["clause_number"],
                "clause_title": r["clause_title_kr"] or "",
                "clause_title_kr": r["clause_title_kr"] or "",
                "clause_topic": r["clause_title_kr"] or "",
                "depth": r["depth"],
                "sort_order": r["sort_order"],
                "source": "standard_clause_masters",
            }
        )
    out.sort(key=lambda x: _clause_sort_key(x["clause_no"]))
    return out


def official_title_map(db: Session, standard_key: str) -> Dict[str, str]:
    return {
        r["clause_no"]: r["clause_title"]
        for r in list_official_clauses(db, standard_key)
        if r.get("clause_title")
    }
