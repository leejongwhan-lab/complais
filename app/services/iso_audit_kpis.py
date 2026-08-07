"""ISO 인증심사 KPI master (chapter-linked) + ESG panel helpers for 심사노트.

- iso_audit_kpi_master ← ComplAIs_인증심사_KPI목록.xlsx (NOT ESG)
- audit_kpi_master     ← process/HLS KPIs (unchanged)
- esg_master_kpis / kpi_master ← ESG panel (unchanged runtime tables)
"""
from __future__ import annotations

import logging
import re
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import load_workbook
from sqlalchemy import text
from sqlalchemy.orm import Session

from app.services.process_group_masters import to_process_standard_code

logger = logging.getLogger(__name__)

BASE_DIR = Path(__file__).resolve().parents[2]
DEFAULT_EXCEL = BASE_DIR / "data" / "ComplAIs_인증심사_KPI목록.xlsx"

# Excel sheet code → process standard_code
_SHEET_TO_STD = {
    "9001": "ISO9001",
    "14001": "ISO14001",
    "45001": "ISO45001",
    "50001": "ISO50001",
    "27001": "ISO27001",
    "27701": "ISO27701",
    "37001": "ISO37001",
    "37301": "ISO37301",
    "22301": "ISO22301",
    "22000": "ISO22000",
    "13485": "ISO13485",
    "42001": "ISO42001",
    "19443": "ISO19443",
}

_STD_TO_ESG_NAME = {
    "ISO9001": "ISO 9001",
    "ISO14001": "ISO 14001",
    "ISO45001": "ISO 45001",
    "ISO50001": "ISO 50001",
    "ISO27001": "ISO 27001",
    "ISO27701": "ISO 27701",
    "ISO37001": "ISO 37001",
    "ISO37301": "ISO 37301",
    "ISO22301": "ISO 22301",
    "ISO22000": "ISO 22000",
    "ISO13485": "ISO 13485",
    "ISO42001": "ISO 42001",
    "ISO19443": "ISO 19443",
}

# Known mis-links in esg_master_kpis (energy rows parked under ISO 9001).
_ESG_REMAP_BY_KPI_ID = {
    17: "ISO 50001",  # 에너지 내부심사 실시
    18: "ISO 50001",  # 에너지 경영검토
}

# Soft domain guard: exclude cross-standard bleed by KPI name keywords.
_ESG_DOMAIN_BLOCK = {
    "ISO9001": re.compile(
        r"에너지|온실가스|GHG|Scope\s*[123]|용수|폐기물|배출권|EnPI|재생에너지",
        re.IGNORECASE,
    ),
    "ISO14001": re.compile(r"제품\s*불량률|납기\s*준수|PPM|품질\s*방침", re.IGNORECASE),
    "ISO45001": re.compile(r"제품\s*불량률|납기\s*준수|PPM|에너지\s*집약", re.IGNORECASE),
}

_DDL = """
CREATE TABLE IF NOT EXISTS iso_audit_kpi_master (
  kpi_id VARCHAR(40) NOT NULL,
  standard_code VARCHAR(30) NOT NULL,
  standard_sheet VARCHAR(20) NULL,
  clause_chapter VARCHAR(40) NOT NULL,
  clause_name VARCHAR(255) NULL,
  kpi_name VARCHAR(255) NOT NULL,
  sort_order SMALLINT NOT NULL DEFAULT 0,
  source VARCHAR(40) NOT NULL DEFAULT 'excel_audit_kpi',
  created_at DATETIME NOT NULL DEFAULT CURRENT_TIMESTAMP,
  PRIMARY KEY (kpi_id),
  KEY ix_iso_audit_kpi_std_chapter (standard_code, clause_chapter),
  KEY ix_iso_audit_kpi_std (standard_code)
) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COLLATE=utf8mb4_unicode_ci
"""


def _table_exists(db: Session, name: str) -> bool:
    row = db.execute(
        text(
            "SELECT 1 FROM information_schema.tables "
            "WHERE table_schema = DATABASE() AND table_name = :t LIMIT 1"
        ),
        {"t": name},
    ).first()
    return bool(row)


def ensure_iso_audit_kpi_table(db: Session) -> None:
    if _table_exists(db, "iso_audit_kpi_master"):
        return
    db.execute(text(_DDL))
    db.commit()


def _cell(v: Any) -> Optional[str]:
    if v is None:
        return None
    s = str(v).strip()
    return s if s else None


def _split_bullets(raw: Optional[str]) -> List[str]:
    if not raw:
        return []
    out: List[str] = []
    for ln in str(raw).splitlines():
        s = ln.strip().lstrip("•·*-").strip()
        if not s:
            continue
        if s.startswith("(") and "전체 적용" in s:
            continue
        out.append(s)
    return out


def _chapter_key(clause_raw: str) -> str:
    s = (clause_raw or "").strip()
    m = re.match(r"^(\d+)", s)
    if m:
        return m.group(1)
    return s[:40]


def parse_iso_audit_kpi_excel(
    excel_path: Path | str = DEFAULT_EXCEL,
) -> List[Dict[str, Any]]:
    excel_path = Path(excel_path)
    if not excel_path.exists():
        raise FileNotFoundError(f"ISO audit KPI Excel not found: {excel_path}")

    wb = load_workbook(excel_path, data_only=True)
    rows_out: List[Dict[str, Any]] = []
    for sheet in wb.sheetnames:
        if sheet == "목차":
            continue
        std = _SHEET_TO_STD.get(str(sheet).strip())
        if not std:
            logger.warning("skip unknown sheet %s", sheet)
            continue
        ws = wb[sheet]
        data = list(ws.iter_rows(values_only=True))
        start = 0
        for i, r in enumerate(data):
            if r and _cell(r[0]) == "조항번호":
                start = i + 1
                break
        chap_seq = 0
        for r in data[start:]:
            if not r:
                continue
            clause_raw = _cell(r[0])
            if not clause_raw:
                continue
            clause_name = _cell(r[1]) if len(r) > 1 else None
            bullets = _split_bullets(_cell(r[2]) if len(r) > 2 else None)
            if not bullets:
                continue
            chap = _chapter_key(clause_raw)
            chap_seq += 1
            for bi, name in enumerate(bullets, 1):
                kpi_id = f"{std}-{chap}-{bi:02d}"
                # specialty chapters may collide across rows — include chap_seq
                if not re.match(r"^\d+$", chap):
                    kpi_id = f"{std}-X{chap_seq}-{bi:02d}"
                rows_out.append(
                    {
                        "kpi_id": kpi_id[:40],
                        "standard_code": std,
                        "standard_sheet": str(sheet).strip(),
                        "clause_chapter": chap,
                        "clause_name": clause_name,
                        "kpi_name": name[:255],
                        "sort_order": chap_seq * 100 + bi,
                        "source": "excel_audit_kpi",
                    }
                )
    return rows_out


def seed_iso_audit_kpis(
    db: Session,
    excel_path: Path | str = DEFAULT_EXCEL,
    *,
    replace: bool = True,
) -> Dict[str, int]:
    """Idempotent seed of iso_audit_kpi_master. Never touches kpi_master/audit_kpi_master."""
    ensure_iso_audit_kpi_table(db)
    rows = parse_iso_audit_kpi_excel(excel_path)
    if replace and _table_exists(db, "iso_audit_kpi_master"):
        db.execute(text("DELETE FROM iso_audit_kpi_master"))
    n = 0
    for r in rows:
        db.execute(
            text(
                "INSERT INTO iso_audit_kpi_master "
                "(kpi_id, standard_code, standard_sheet, clause_chapter, "
                "clause_name, kpi_name, sort_order, source) "
                "VALUES (:id, :std, :sheet, :chap, :cname, :name, :ord, :src) "
                "ON DUPLICATE KEY UPDATE "
                "standard_code=VALUES(standard_code), "
                "standard_sheet=VALUES(standard_sheet), "
                "clause_chapter=VALUES(clause_chapter), "
                "clause_name=VALUES(clause_name), "
                "kpi_name=VALUES(kpi_name), "
                "sort_order=VALUES(sort_order), "
                "source=VALUES(source)"
            ),
            {
                "id": r["kpi_id"],
                "std": r["standard_code"],
                "sheet": r["standard_sheet"],
                "chap": r["clause_chapter"],
                "cname": r["clause_name"],
                "name": r["kpi_name"],
                "ord": r["sort_order"],
                "src": r["source"],
            },
        )
        n += 1
    db.commit()
    # safety: companies / CBs untouched — return counts for verify
    companies = 0
    cbs = 0
    try:
        if _table_exists(db, "companies"):
            companies = int(db.execute(text("SELECT COUNT(*) FROM companies")).scalar() or 0)
        if _table_exists(db, "certification_bodies"):
            cbs = int(
                db.execute(text("SELECT COUNT(*) FROM certification_bodies")).scalar() or 0
            )
    except Exception:
        pass
    logger.info(
        "iso_audit_kpi_master seeded n=%s companies=%s cbs=%s", n, companies, cbs
    )
    return {
        "iso_audit_kpi_master": n,
        "companies": companies,
        "certification_bodies": cbs,
    }


def _major_chapter(clause_no: str) -> Optional[str]:
    m = re.match(r"^(\d+)", str(clause_no or "").strip())
    return m.group(1) if m else None


def list_iso_audit_kpis_for_clause(
    db: Session,
    *,
    standard_code: Optional[str],
    clause_no: str,
) -> List[Dict[str, str]]:
    ensure_iso_audit_kpi_table(db)
    if not _table_exists(db, "iso_audit_kpi_master"):
        return []
    pg = to_process_standard_code(standard_code) or (standard_code or "").strip()
    if not pg:
        return []
    chap = _major_chapter(clause_no)
    sql = (
        "SELECT kpi_id, kpi_name, clause_chapter, clause_name "
        "FROM iso_audit_kpi_master WHERE standard_code = :std "
    )
    params: Dict[str, Any] = {"std": pg}
    if chap:
        # numeric chapter match + specialty (non-digit) chapters for the standard
        sql += (
            "AND (clause_chapter = :chap OR clause_chapter REGEXP '[^0-9]') "
            "ORDER BY sort_order, kpi_id"
        )
        params["chap"] = chap
    else:
        sql += "ORDER BY sort_order, kpi_id"
    rows = db.execute(text(sql), params).mappings().all()
    out: List[Dict[str, str]] = []
    for r in rows:
        kid = str(r["kpi_id"])
        kn = str(r["kpi_name"] or kid)
        out.append(
            {
                "kpi_id": kid,
                "kpi_name": kn,
                "key": kid,
                "label": kn,
                "source": "iso_audit",
                "kpi_kind": "iso",
            }
        )
    return out


def _parse_clause_range(detail: str) -> Optional[Tuple[int, int]]:
    """Parse '6~9 (…)' / '8~9' / '5~8' → (lo, hi). Returns None if unparseable."""
    s = (detail or "").strip()
    if not s or s == "-":
        return None
    m = re.search(r"(\d+)\s*[~～\-–—]\s*(\d+)", s)
    if m:
        return int(m.group(1)), int(m.group(2))
    m = re.match(r"^(\d+)", s)
    if m:
        n = int(m.group(1))
        return n, n
    return None


def _iso_num_from_code(pg: str) -> Optional[str]:
    m = re.search(r"(\d{4,5})", pg or "")
    return m.group(1) if m else None


def _resolve_esg_standard_names(db: Session, pg: str) -> List[str]:
    """Map process standard_code → managed_standard_name values present in master."""
    preferred = _STD_TO_ESG_NAME.get(pg or "", "")
    names: List[str] = []
    if preferred:
        names.append(preferred)
    num = _iso_num_from_code(pg)
    if not num or not _table_exists(db, "esg_master_kpis"):
        return names
    try:
        rows = db.execute(
            text(
                "SELECT DISTINCT managed_standard_name FROM esg_master_kpis "
                "WHERE managed_standard_name LIKE :pat "
                "ORDER BY managed_standard_name"
            ),
            {"pat": f"%{num}%"},
        ).mappings().all()
        for r in rows:
            n = (r.get("managed_standard_name") or "").strip()
            if n and n not in names and n not in {"기타/확인필요", "공통"}:
                names.append(n)
    except Exception:
        try:
            db.rollback()
        except Exception:
            pass
    return names


def ensure_esg_standard_kpi_links(db: Session) -> Dict[str, int]:
    """Soft-fix known mis-linked esg_master_kpis rows (no DROP). Idempotent."""
    fixed = 0
    if not _table_exists(db, "esg_master_kpis"):
        return {"fixed": 0}
    for kid, target in _ESG_REMAP_BY_KPI_ID.items():
        try:
            res = db.execute(
                text(
                    "UPDATE esg_master_kpis SET managed_standard_name = :tgt "
                    "WHERE kpi_id = :kid AND managed_standard_name <> :tgt"
                ),
                {"tgt": target, "kid": kid},
            )
            fixed += int(res.rowcount or 0)
        except Exception:
            try:
                db.rollback()
            except Exception:
                pass
            logger.exception("esg kpi remap failed for kpi_id=%s", kid)
    # Energy-named rows still parked on ISO 9001 → ISO 50001
    try:
        res = db.execute(
            text(
                "UPDATE esg_master_kpis SET managed_standard_name = 'ISO 50001' "
                "WHERE managed_standard_name = 'ISO 9001' "
                "AND is_iso_auditable = 1 "
                "AND (kpi_name LIKE '%에너지%' OR kpi_name LIKE '%EnPI%' "
                "OR kpi_name LIKE '%재생에너지%')"
            )
        )
        fixed += int(res.rowcount or 0)
    except Exception:
        try:
            db.rollback()
        except Exception:
            pass
        logger.exception("esg energy→50001 remap failed")
    if fixed:
        try:
            db.commit()
        except Exception:
            try:
                db.rollback()
            except Exception:
                pass
    return {"fixed": fixed}


def list_esg_kpis_for_clause(
    db: Session,
    *,
    standard_code: Optional[str],
    clause_no: str,
    limit: int = 40,
) -> List[Dict[str, str]]:
    """ESG panel — esg_master_kpis aligned to selected standard + clause chapter."""
    pg = to_process_standard_code(standard_code) or (standard_code or "").strip()
    chap = _major_chapter(clause_no)
    chap_n = int(chap) if chap and chap.isdigit() else None
    block_re = _ESG_DOMAIN_BLOCK.get(pg or "")

    if _table_exists(db, "esg_master_kpis"):
        esg_names = _resolve_esg_standard_names(db, pg or "")
        rows: List[Any] = []
        try:
            if esg_names:
                # Exact preferred name first; also accept fuzzy ISO-number matches
                placeholders = ", ".join(f":n{i}" for i in range(len(esg_names)))
                params: Dict[str, Any] = {f"n{i}": n for i, n in enumerate(esg_names)}
                rows = db.execute(
                    text(
                        "SELECT kpi_id, kpi_name, esg_category, managed_standard_name, "
                        "iso_clause_detail, unit_format "
                        "FROM esg_master_kpis "
                        "WHERE is_iso_auditable = 1 "
                        f"AND managed_standard_name IN ({placeholders}) "
                        "ORDER BY kpi_id"
                    ),
                    params,
                ).mappings().all()
        except Exception:
            try:
                db.rollback()
            except Exception:
                pass
            rows = []
        matched: List[Dict[str, str]] = []
        preferred = _STD_TO_ESG_NAME.get(pg or "", "")
        for r in rows:
            mname = str(r.get("managed_standard_name") or "").strip()
            # Prefer exact standard label when multiple fuzzy hits exist
            if preferred and mname != preferred and len(esg_names) > 1:
                # still allow if only fuzzy variants exist for this ISO number
                if any(n == preferred for n in esg_names) and mname != preferred:
                    continue
            kn_raw = str(r.get("kpi_name") or "")
            if block_re and block_re.search(kn_raw):
                continue
            detail = str(r.get("iso_clause_detail") or "")
            rng = _parse_clause_range(detail)
            if chap_n is not None and rng is not None:
                if not (rng[0] <= chap_n <= rng[1]):
                    continue
            elif chap_n is not None and rng is None:
                # no clause chapter link — skip for specific chapter view
                continue
            kid = f"esg:{r['kpi_id']}"
            cat = r.get("esg_category") or ""
            kn = kn_raw or kid
            label = f"[{cat}] {kn}" if cat else kn
            matched.append(
                {
                    "kpi_id": kid,
                    "kpi_name": label,
                    "key": kid,
                    "label": label,
                    "source": "esg_master",
                    "kpi_kind": "esg",
                    "managed_standard_name": mname,
                    "iso_clause_detail": detail,
                }
            )
            if len(matched) >= limit:
                break
        if matched:
            return matched

    # fallback: kpi_master (ESG runtime) — filter by applicable_stds / iso_clause
    if _table_exists(db, "kpi_master") and chap_n is not None:
        num = _iso_num_from_code(pg or "")
        try:
            rows = db.execute(
                text(
                    "SELECT id, name_kr, category_esg, iso_clause, applicable_stds "
                    "FROM kpi_master "
                    "WHERE is_active = 1 AND iso_clause IS NOT NULL "
                    "AND TRIM(iso_clause) <> '' "
                    "ORDER BY sort_order, id LIMIT :lim"
                ),
                {"lim": max(limit * 3, 60)},
            ).mappings().all()
        except Exception:
            try:
                db.rollback()
            except Exception:
                pass
            return []
        out: List[Dict[str, str]] = []
        for r in rows:
            appl = str(r.get("applicable_stds") or "")
            if num and appl and num not in appl and "공통" not in appl:
                continue
            kn_raw = str(r.get("name_kr") or "")
            if block_re and block_re.search(kn_raw):
                continue
            detail = str(r.get("iso_clause") or "")
            rng = _parse_clause_range(detail)
            if rng and not (rng[0] <= chap_n <= rng[1]):
                continue
            if not rng and not detail.startswith(str(chap_n)):
                continue
            kid = f"kpi:{r['id']}"
            cat = r.get("category_esg") or ""
            kn = kn_raw or kid
            label = f"[{cat}] {kn}" if cat else kn
            out.append(
                {
                    "kpi_id": kid,
                    "kpi_name": label,
                    "key": kid,
                    "label": label,
                    "source": "kpi_master",
                    "kpi_kind": "esg",
                }
            )
            if len(out) >= limit:
                break
        return out
    return []


def enrich_clause_dual_kpis(
    db: Session,
    clause: Dict[str, Any],
) -> Dict[str, Any]:
    """Attach iso_audit_kpis + esg_kpis; merge ISO panel = audit_kpi + excel chapter KPIs."""
    std = clause.get("standard_code")
    cno = clause.get("clause_no") or ""
    excel_kpis = list_iso_audit_kpis_for_clause(
        db, standard_code=std, clause_no=cno
    )
    # existing HLS audit_kpi_master rows already on default_kpis
    hls_kpis = []
    for k in clause.get("default_kpis") or []:
        kid = (k.get("kpi_id") or k.get("key") or "").strip()
        kn = (k.get("kpi_name") or k.get("label") or kid).strip()
        if not kid:
            continue
        hls_kpis.append(
            {
                "kpi_id": kid,
                "kpi_name": kn,
                "key": kid,
                "label": kn,
                "source": "audit_kpi",
                "kpi_kind": "iso",
            }
        )
    # de-dupe by kpi_id — excel first, then hls
    seen = set()
    iso_panel: List[Dict[str, str]] = []
    for k in excel_kpis + hls_kpis:
        if k["kpi_id"] in seen:
            continue
        seen.add(k["kpi_id"])
        iso_panel.append(k)
    clause["iso_audit_kpis"] = iso_panel
    clause["esg_kpis"] = list_esg_kpis_for_clause(
        db, standard_code=std, clause_no=cno
    )
    return clause
