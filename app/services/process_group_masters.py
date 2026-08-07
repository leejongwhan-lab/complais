"""Process-group / HLS / standard-map / audit KPI masters.

Excel seed: data/ISO_심사체크포인트_DB시드_프로세스그룹.xlsx

Table names (conflict-safe):
- process_group_master          ← Excel PROCESS_GROUP_MASTER
- standard_master               ← Excel STANDARD_MASTER (≠ platform standard_masters)
- hls_master                    ← Excel HLS_MASTER
- process_group_hls_map         ← Excel PROCESS_GROUP_HLS_MAP
- standard_process_clause_map   ← Excel STANDARD_PROCESS_CLAUSE_MAP
- standard_clause_map           ← Excel STANDARD_CLAUSE_MAP
- audit_kpi_master              ← Excel KPI_MASTER (kpi_master is ESG; do not reuse)
"""
from __future__ import annotations

import logging
import re
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import load_workbook
from sqlalchemy import text
from sqlalchemy.orm import Session

logger = logging.getLogger(__name__)

BASE_DIR = Path(__file__).resolve().parents[2]
DEFAULT_EXCEL = BASE_DIR / "data" / "ISO_심사체크포인트_DB시드_프로세스그룹.xlsx"

# Excel KPI_MASTER typos / aliases → canonical hls_master keys
_HLS_CODE_NORMALIZE = {
    "8.2~8.10": "8.6~8.10",
}

# Common HLS clause titles (Korean) — used when iso_clauses / clause_topic absent
_HLS_TITLE_KO: Dict[str, str] = {
    "4.1": "조직과 그 상황의 이해",
    "4.2": "이해관계자의 니즈와 기대 이해",
    "4.3": "경영시스템의 적용범위 결정",
    "4.4": "경영시스템",
    "5.1": "리더십과 의지표명",
    "5.2": "방침",
    "5.3": "조직의 역할, 책임 및 권한",
    "6.1": "리스크와 기회를 다루는 조치",
    "6.2": "경영시스템 목표와 이를 달성하기 위한 기획",
    "6.3": "변경의 기획",
    "7.1": "자원",
    "7.2": "역량",
    "7.3": "인식",
    "7.4": "의사소통",
    "7.5": "문서화된 정보",
    "8.1": "운용 기획 및 관리",
    "8.2": "요구사항(제품·서비스 / 표준별)",
    "8.3": "설계 및 개발(표준별)",
    "8.4": "외부에서 제공되는 프로세스·제품·서비스의 관리",
    "8.5": "생산 및 서비스 제공(표준별)",
    "8.6": "제품 및 서비스의 불출(표준별)",
    "8.7": "부적합 출력물의 관리(표준별)",
    "8.6~8.10": "운용 세부단계(표준별 8.6~8.10)",
    "9.1": "모니터링, 측정, 분석 및 평가",
    "9.2": "내부심사",
    "9.3": "경영검토",
    "10.1": "일반(개선)",
    "10.2": "부적합 및 시정조치",
    "10.3": "지속적 개선",
}

# platform standard_key / family → Excel STANDARD_MASTER.standard_code
_FAMILY_TO_PG_CODE = {
    "QMS": "ISO9001",
    "EMS": "ISO14001",
    "OHSMS": "ISO45001",
    "ISMS": "ISO27001",
    "ABMS": "ISO37001",
    "CMS": "ISO37301",
    "EnMS": "ISO50001",
    "FSMS": "ISO22000",
    "NSMS": "ISO19443",
    "PIMS": "ISO27701",
    "AIMS": "ISO42001",
}

# CREATE TABLE IF NOT EXISTS — mirrors alembic f4a5b6c7d8e9 (runtime safety net)
_DDL = [
    """
CREATE TABLE IF NOT EXISTS process_group_master (
  process_group_id VARCHAR(10) NOT NULL,
  process_group_name VARCHAR(100) NOT NULL,
  hls_scope_desc VARCHAR(255) NULL,
  PRIMARY KEY (process_group_id)
) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COLLATE=utf8mb4_unicode_ci
""",
    """
CREATE TABLE IF NOT EXISTS standard_master (
  standard_code VARCHAR(30) NOT NULL,
  standard_name VARCHAR(200) NOT NULL,
  hls_adopted VARCHAR(10) NOT NULL DEFAULT 'Y',
  native_structure_note TEXT NULL,
  PRIMARY KEY (standard_code)
) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COLLATE=utf8mb4_unicode_ci
""",
    """
CREATE TABLE IF NOT EXISTS hls_master (
  hls_code VARCHAR(20) NOT NULL,
  checkpoints_summary TEXT NULL,
  PRIMARY KEY (hls_code)
) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COLLATE=utf8mb4_unicode_ci
""",
    """
CREATE TABLE IF NOT EXISTS process_group_hls_map (
  id INT NOT NULL AUTO_INCREMENT,
  process_group_id VARCHAR(10) NOT NULL,
  hls_code VARCHAR(20) NOT NULL,
  PRIMARY KEY (id),
  UNIQUE KEY uq_process_group_hls (process_group_id, hls_code),
  CONSTRAINT fk_pg_hls_map_pg FOREIGN KEY (process_group_id)
    REFERENCES process_group_master (process_group_id),
  CONSTRAINT fk_pg_hls_map_hls FOREIGN KEY (hls_code)
    REFERENCES hls_master (hls_code)
) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COLLATE=utf8mb4_unicode_ci
""",
    """
CREATE TABLE IF NOT EXISTS standard_process_clause_map (
  id INT NOT NULL AUTO_INCREMENT,
  standard_code VARCHAR(30) NOT NULL,
  process_group_id VARCHAR(10) NOT NULL,
  actual_clause_no VARCHAR(30) NOT NULL,
  clause_topic VARCHAR(255) NULL,
  guide_note TEXT NULL,
  PRIMARY KEY (id),
  UNIQUE KEY uq_std_proc_clause (standard_code, process_group_id, actual_clause_no),
  KEY ix_std_proc_clause_std (standard_code),
  KEY ix_std_proc_clause_pg (process_group_id),
  CONSTRAINT fk_std_proc_clause_std FOREIGN KEY (standard_code)
    REFERENCES standard_master (standard_code),
  CONSTRAINT fk_std_proc_clause_pg FOREIGN KEY (process_group_id)
    REFERENCES process_group_master (process_group_id)
) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COLLATE=utf8mb4_unicode_ci
""",
    """
CREATE TABLE IF NOT EXISTS standard_clause_map (
  id INT NOT NULL AUTO_INCREMENT,
  standard_code VARCHAR(30) NOT NULL,
  hls_code VARCHAR(20) NOT NULL,
  actual_clause_no VARCHAR(30) NULL,
  status VARCHAR(20) NOT NULL DEFAULT 'DIRECT',
  integrated_into_hls_code VARCHAR(20) NULL,
  guide_note TEXT NULL,
  PRIMARY KEY (id),
  UNIQUE KEY uq_std_clause_hls (standard_code, hls_code),
  KEY ix_std_clause_map_std (standard_code),
  CONSTRAINT fk_std_clause_map_std FOREIGN KEY (standard_code)
    REFERENCES standard_master (standard_code),
  CONSTRAINT fk_std_clause_map_hls FOREIGN KEY (hls_code)
    REFERENCES hls_master (hls_code)
) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COLLATE=utf8mb4_unicode_ci
""",
    """
CREATE TABLE IF NOT EXISTS audit_kpi_master (
  kpi_id VARCHAR(30) NOT NULL,
  hls_code VARCHAR(20) NOT NULL,
  standard_code VARCHAR(30) NOT NULL DEFAULT 'COMMON',
  kpi_name VARCHAR(255) NOT NULL,
  kpi_type ENUM('RATIO','COUNT','PERIOD','TREND') NOT NULL,
  formula TEXT NULL,
  unit VARCHAR(50) NULL,
  PRIMARY KEY (kpi_id),
  KEY ix_audit_kpi_hls (hls_code),
  KEY ix_audit_kpi_std (standard_code),
  CONSTRAINT fk_audit_kpi_hls FOREIGN KEY (hls_code)
    REFERENCES hls_master (hls_code)
) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COLLATE=utf8mb4_unicode_ci
""",
]

# Child → parent order for safe replace of seed rows (our tables only)
_CLEAR_ORDER = [
    "audit_kpi_master",
    "standard_clause_map",
    "standard_process_clause_map",
    "process_group_hls_map",
    "hls_master",
    "standard_master",
    "process_group_master",
]

_KPI_TYPES = {"RATIO", "COUNT", "PERIOD", "TREND"}


def _table_exists(db: Session, name: str) -> bool:
    row = db.execute(
        text(
            "SELECT 1 FROM information_schema.tables "
            "WHERE table_schema = DATABASE() AND table_name = :t LIMIT 1"
        ),
        {"t": name},
    ).first()
    return bool(row)


def _cell(v: Any) -> Optional[str]:
    if v is None:
        return None
    s = str(v).strip()
    return s if s else None


def _normalize_hls(code: Optional[str]) -> Optional[str]:
    if not code:
        return None
    return _HLS_CODE_NORMALIZE.get(code, code)


def to_process_standard_code(raw: Optional[str]) -> Optional[str]:
    """Map platform standard_key / display / ISO code → Excel standard_code (ISO9001…)."""
    if raw is None:
        return None
    s = str(raw).strip()
    if not s:
        return None
    compact = re.sub(r"[\s/._-]+", "", s.upper())
    # Already ISO9001-style
    if re.fullmatch(r"ISO\d{4,5}", compact) or re.fullmatch(r"ISOIEC\d{4,5}", compact):
        digits = re.search(r"(\d{4,5})", compact)
        return f"ISO{digits.group(1)}" if digits else None
    # QMS_2015 / EMS_2015
    fam = s.split("_")[0] if "_" in s else s
    if fam in _FAMILY_TO_PG_CODE:
        return _FAMILY_TO_PG_CODE[fam]
    upper = fam.upper()
    for k, v in _FAMILY_TO_PG_CODE.items():
        if upper == k.upper():
            return v
    # ISO 9001:2015 / ISO/IEC 27001
    m = re.search(r"(\d{4,5})", s)
    if m and ("ISO" in compact or "IEC" in compact):
        return f"ISO{m.group(1)}"
    return None


def _clause_sort_key(clause_no: str) -> Tuple:
    parts = []
    for p in re.split(r"[.\-/~]", str(clause_no)):
        if p.isdigit():
            parts.append((0, int(p)))
        else:
            parts.append((1, p))
    return tuple(parts) if parts else ((1, str(clause_no)),)


def _split_checkpoints(summary: Optional[str]) -> List[Dict[str, str]]:
    if not summary:
        return []
    out = []
    for part in re.split(r"[/\n|;]+", str(summary)):
        title = part.strip()
        if title:
            out.append({"title": title, "hint": ""})
    return out


_EMS_DEEP_RE = re.compile(
    r"환경심층|TCFD|기후관련재무|Scope\s*[123]|GHG\s*Protocol",
    re.IGNORECASE,
)


def _is_ems_standard(standard_code: Optional[str]) -> bool:
    s = (standard_code or "").upper()
    return "14001" in s or s.startswith("EMS") or s == "ISO14001"


def _filter_checkpoints_for_standard(
    cps: List[Dict[str, str]], standard_code: Optional[str]
) -> List[Dict[str, str]]:
    """Drop EMS 환경심층 / TCFD fragments leaking into non-EMS standards."""
    if not cps or _is_ems_standard(standard_code):
        return cps
    out = []
    for c in cps:
        blob = f"{c.get('title') or ''} {c.get('hint') or ''}"
        if _EMS_DEEP_RE.search(blob):
            continue
        out.append(c)
    return out


def _clean_clause_title(clause_no: Optional[str], title: Optional[str]) -> str:
    """Return title without leading clause_no duplicate (e.g. '6.1c 기후…' → '기후…')."""
    s = (title or "").strip()
    if not s:
        return ""
    cno = (clause_no or "").strip()
    if not cno:
        return s
    # strip exact / case-insensitive leading number (+ optional punctuation)
    pat = re.compile(
        r"^\s*" + re.escape(cno) + r"(?:\s*[:.\-)\]]\s*|\s+)",
        re.IGNORECASE,
    )
    s2 = pat.sub("", s, count=1).strip()
    if s2:
        return s2
    # also strip if title starts with same digits ignoring letter case (6.1C vs 6.1c)
    pat2 = re.compile(r"^\s*[0-9]+(?:\.[0-9A-Za-z]+)*(?:\s*[:.\-)\]]\s*|\s+)")
    m = pat2.match(s)
    if m and cno.lower().rstrip("c") in m.group(0).lower():
        return s[m.end():].strip() or s
    return s


def _hls_matches_clause(kpi_hls: str, clause_hls: str, clause_no: str) -> bool:
    if not kpi_hls:
        return False
    if kpi_hls == clause_hls or kpi_hls == clause_no:
        return True
    if "~" in kpi_hls:
        a, b = kpi_hls.split("~", 1)
        a, b = a.strip(), b.strip()
        for cand in (clause_hls, clause_no):
            if not cand or "~" in cand or "/" in cand:
                # multi-clause labels like 7.1.5/7.1.6 — check primary
                primary = cand.split("/")[0].strip() if cand else ""
                cand = primary
            try:
                if _clause_sort_key(a) <= _clause_sort_key(cand) <= _clause_sort_key(b):
                    return True
            except Exception:
                pass
    return False



def ensure_ncr_extra_columns(db: Session) -> None:
    """Additive columns for 부적합 보고서 modal fields (no DROP)."""
    if not _table_exists(db, "audit_note_ncr"):
        return
    cols = {
        "request_date": "DATE NULL",
        "reported_at": "DATE NULL",
        "auditor_name": "VARCHAR(100) NULL",
        "dept": "VARCHAR(200) NULL",
        "esg_tags": "LONGTEXT NULL",
    }
    for col, typ in cols.items():
        row = db.execute(
            text(
                "SELECT 1 FROM information_schema.columns "
                "WHERE table_schema=DATABASE() AND table_name='audit_note_ncr' "
                "AND column_name=:c LIMIT 1"
            ),
            {"c": col},
        ).first()
        if row:
            continue
        try:
            db.execute(text(f"ALTER TABLE audit_note_ncr ADD COLUMN `{col}` {typ}"))
        except Exception as exc:
            logger.warning("add column audit_note_ncr.%s failed: %s", col, exc)
    # promote root_cause to LONGTEXT if present
    row = db.execute(
        text(
            "SELECT DATA_TYPE FROM information_schema.columns "
            "WHERE table_schema=DATABASE() AND table_name='audit_note_ncr' "
            "AND column_name='root_cause'"
        )
    ).first()
    if row and str(row[0]).lower() != "longtext":
        try:
            db.execute(text("ALTER TABLE audit_note_ncr MODIFY COLUMN root_cause LONGTEXT NULL"))
        except Exception as exc:
            logger.warning("LONGTEXT root_cause failed: %s", exc)
    db.commit()


def ensure_audit_note_longtext(db: Session) -> None:
    """Promote note/evidence columns to LONGTEXT (internal storage, not public)."""
    alters = [
        ("audit_note_clauses", "finding", "LONGTEXT NULL"),
        ("audit_note_clauses", "evidence", "LONGTEXT NULL"),
        ("audit_note_clauses", "kpi_json", "LONGTEXT NULL"),
        ("audit_note_ncr", "description", "LONGTEXT NULL"),
        ("audit_note_ncr", "requirement", "LONGTEXT NULL"),
        ("audit_note_ncr", "evidence", "LONGTEXT NULL"),
        ("audit_notes", "content", "LONGTEXT NULL"),
        ("audit_notes", "summary", "LONGTEXT NULL"),
        ("hls_master", "checkpoints_summary", "LONGTEXT NULL"),
    ]
    for table, col, typ in alters:
        if not _table_exists(db, table):
            continue
        row = db.execute(
            text(
                "SELECT DATA_TYPE FROM information_schema.columns "
                "WHERE table_schema=DATABASE() AND table_name=:t AND column_name=:c"
            ),
            {"t": table, "c": col},
        ).first()
        if not row:
            continue
        if str(row[0]).lower() != "longtext":
            try:
                db.execute(text(f"ALTER TABLE `{table}` MODIFY COLUMN `{col}` {typ}"))
            except Exception as exc:
                logger.warning("LONGTEXT alter %s.%s failed: %s", table, col, exc)

    # Additive session columns (runtime safety net if alembic not yet applied)
    additive = [
        (
            "audit_notes",
            "interview_json",
            "LONGTEXT NULL COMMENT '면담(interview) 기록 JSON'",
        ),
        (
            "audit_notes",
            "note_method",
            "VARCHAR(20) NULL DEFAULT 'process' "
            "COMMENT '심사방식: clause|process'",
        ),
        (
            "audit_note_clauses",
            "audit_method",
            "VARCHAR(20) NULL COMMENT '작성 시 심사방식 clause|process'",
        ),
        (
            "audit_note_clauses",
            "standard_code",
            "VARCHAR(30) NULL COMMENT 'standard_master.standard_code'",
        ),
        (
            "audit_note_clauses",
            "process_group_id",
            "VARCHAR(10) NULL COMMENT 'process_group_master.process_group_id'",
        ),
        (
            "audit_note_clauses",
            "hls_code",
            "VARCHAR(20) NULL COMMENT 'hls_master.hls_code'",
        ),
        (
            "audit_note_clauses",
            "clause_topic",
            "VARCHAR(255) NULL COMMENT 'clause_topic from master'",
        ),
    ]
    for table, col, typ in additive:
        if not _table_exists(db, table):
            continue
        row = db.execute(
            text(
                "SELECT 1 FROM information_schema.columns "
                "WHERE table_schema=DATABASE() AND table_name=:t AND column_name=:c"
            ),
            {"t": table, "c": col},
        ).first()
        if row:
            continue
        try:
            db.execute(text(f"ALTER TABLE `{table}` ADD COLUMN `{col}` {typ}"))
        except Exception as exc:
            logger.warning("ADD %s.%s failed: %s", table, col, exc)
    db.commit()


def ensure_process_group_tables(db: Session) -> None:
    for sql in _DDL:
        db.execute(text(sql))
    ensure_ncr_extra_columns(db)
    db.commit()


def _sheet_rows(wb, name: str) -> List[tuple]:
    if name not in wb.sheetnames:
        return []
    ws = wb[name]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    out = []
    for r in rows[1:]:
        if r and r[0] is not None and str(r[0]).strip():
            out.append(r)
    return out


def seed_process_group_masters(
    db: Session,
    excel_path: Optional[Path] = None,
    *,
    replace: bool = True,
) -> Dict[str, int]:
    """Load Excel sheets into master tables. Returns row counts per table."""
    path = Path(excel_path) if excel_path else DEFAULT_EXCEL
    if not path.exists():
        raise FileNotFoundError(f"Excel seed not found: {path}")

    ensure_process_group_tables(db)
    ensure_audit_note_longtext(db)
    wb = load_workbook(path, data_only=True)

    if replace:
        # Only our seed tables — never touch kpi_master / standard_masters / companies
        for t in _CLEAR_ORDER:
            if _table_exists(db, t):
                db.execute(text(f"DELETE FROM `{t}`"))
        db.commit()

    counts: Dict[str, int] = {}

    # 1) process_group_master
    n = 0
    for r in _sheet_rows(wb, "PROCESS_GROUP_MASTER"):
        db.execute(
            text(
                "INSERT INTO process_group_master "
                "(process_group_id, process_group_name, hls_scope_desc) "
                "VALUES (:id, :name, :scope) "
                "ON DUPLICATE KEY UPDATE "
                "process_group_name=VALUES(process_group_name), "
                "hls_scope_desc=VALUES(hls_scope_desc)"
            ),
            {
                "id": _cell(r[0]),
                "name": _cell(r[1]) or "",
                "scope": _cell(r[2]) if len(r) > 2 else None,
            },
        )
        n += 1
    counts["process_group_master"] = n

    # 2) standard_master
    n = 0
    for r in _sheet_rows(wb, "STANDARD_MASTER"):
        db.execute(
            text(
                "INSERT INTO standard_master "
                "(standard_code, standard_name, hls_adopted, native_structure_note) "
                "VALUES (:code, :name, :hls, :note) "
                "ON DUPLICATE KEY UPDATE "
                "standard_name=VALUES(standard_name), "
                "hls_adopted=VALUES(hls_adopted), "
                "native_structure_note=VALUES(native_structure_note)"
            ),
            {
                "code": _cell(r[0]),
                "name": _cell(r[1]) or "",
                "hls": _cell(r[2]) or "Y",
                "note": _cell(r[3]) if len(r) > 3 else None,
            },
        )
        n += 1
    counts["standard_master"] = n

    # 3) hls_master (+ stub codes referenced by KPI/maps but absent from HLS sheet)
    n = 0
    known_hls: set = set()
    for r in _sheet_rows(wb, "HLS_MASTER"):
        code = _cell(r[0])
        if not code:
            continue
        known_hls.add(code)
        db.execute(
            text(
                "INSERT INTO hls_master (hls_code, checkpoints_summary) "
                "VALUES (:code, :summary) "
                "ON DUPLICATE KEY UPDATE "
                "checkpoints_summary=VALUES(checkpoints_summary)"
            ),
            {"code": code, "summary": _cell(r[1]) if len(r) > 1 else None},
        )
        n += 1

    # Normalize KPI typos (8.2~8.10 → 8.6~8.10); only stub truly missing HLS keys
    extra_hls: set = set()
    for sheet, col in (
        ("KPI_MASTER", 1),
        ("PROCESS_GROUP_HLS_MAP", 1),
        ("STANDARD_CLAUSE_MAP", 1),
    ):
        for r in _sheet_rows(wb, sheet):
            code = _normalize_hls(_cell(r[col]) if len(r) > col else None)
            if code and code not in known_hls:
                extra_hls.add(code)
    for code in sorted(extra_hls):
        db.execute(
            text(
                "INSERT INTO hls_master (hls_code, checkpoints_summary) "
                "VALUES (:code, :summary) "
                "ON DUPLICATE KEY UPDATE hls_code=hls_code"
            ),
            {
                "code": code,
                "summary": "(seed stub) referenced by KPI/map; not in HLS_MASTER sheet",
            },
        )
        n += 1
        known_hls.add(code)
    # Drop obsolete stub if a prior seed inserted 8.2~8.10
    if "8.2~8.10" not in known_hls:
        try:
            db.execute(text("DELETE FROM hls_master WHERE hls_code = '8.2~8.10'"))
        except Exception:
            pass
    counts["hls_master"] = n

    # 4) process_group_hls_map
    n = 0
    for r in _sheet_rows(wb, "PROCESS_GROUP_HLS_MAP"):
        db.execute(
            text(
                "INSERT INTO process_group_hls_map (process_group_id, hls_code) "
                "VALUES (:pg, :hls) "
                "ON DUPLICATE KEY UPDATE hls_code=VALUES(hls_code)"
            ),
            {"pg": _cell(r[0]), "hls": _normalize_hls(_cell(r[1]))},
        )
        n += 1
    counts["process_group_hls_map"] = n

    # 5) standard_process_clause_map
    n = 0
    for r in _sheet_rows(wb, "STANDARD_PROCESS_CLAUSE_MAP"):
        db.execute(
            text(
                "INSERT INTO standard_process_clause_map "
                "(standard_code, process_group_id, actual_clause_no, clause_topic, guide_note) "
                "VALUES (:std, :pg, :clause, :topic, :note) "
                "ON DUPLICATE KEY UPDATE "
                "clause_topic=VALUES(clause_topic), guide_note=VALUES(guide_note)"
            ),
            {
                "std": _cell(r[0]),
                "pg": _cell(r[1]),
                "clause": _cell(r[2]) or "",
                "topic": _cell(r[3]) if len(r) > 3 else None,
                "note": _cell(r[4]) if len(r) > 4 else None,
            },
        )
        n += 1
    counts["standard_process_clause_map"] = n

    # 6) standard_clause_map (common-clause status map)
    n = 0
    for r in _sheet_rows(wb, "STANDARD_CLAUSE_MAP"):
        db.execute(
            text(
                "INSERT INTO standard_clause_map "
                "(standard_code, hls_code, actual_clause_no, status, "
                " integrated_into_hls_code, guide_note) "
                "VALUES (:std, :hls, :clause, :status, :into, :note) "
                "ON DUPLICATE KEY UPDATE "
                "actual_clause_no=VALUES(actual_clause_no), "
                "status=VALUES(status), "
                "integrated_into_hls_code=VALUES(integrated_into_hls_code), "
                "guide_note=VALUES(guide_note)"
            ),
            {
                "std": _cell(r[0]),
                "hls": _normalize_hls(_cell(r[1])),
                "clause": _cell(r[2]) if len(r) > 2 else None,
                "status": _cell(r[3]) if len(r) > 3 else "DIRECT",
                "into": _cell(r[4]) if len(r) > 4 else None,
                "note": _cell(r[5]) if len(r) > 5 else None,
            },
        )
        n += 1
    counts["standard_clause_map"] = n

    # 7) audit_kpi_master ← Excel KPI_MASTER
    n = 0
    for r in _sheet_rows(wb, "KPI_MASTER"):
        kpi_type = (_cell(r[4]) if len(r) > 4 else None) or "RATIO"
        if kpi_type not in _KPI_TYPES:
            logger.warning("skip kpi %s unknown type %s", r[0], kpi_type)
            continue
        db.execute(
            text(
                "INSERT INTO audit_kpi_master "
                "(kpi_id, hls_code, standard_code, kpi_name, kpi_type, formula, unit) "
                "VALUES (:id, :hls, :std, :name, :typ, :formula, :unit) "
                "ON DUPLICATE KEY UPDATE "
                "hls_code=VALUES(hls_code), standard_code=VALUES(standard_code), "
                "kpi_name=VALUES(kpi_name), kpi_type=VALUES(kpi_type), "
                "formula=VALUES(formula), unit=VALUES(unit)"
            ),
            {
                "id": _cell(r[0]),
                "hls": _normalize_hls(_cell(r[1])),
                "std": _cell(r[2]) or "COMMON",
                "name": _cell(r[3]) or "",
                "typ": kpi_type,
                "formula": _cell(r[5]) if len(r) > 5 else None,
                "unit": _cell(r[6]) if len(r) > 6 else None,
            },
        )
        n += 1
    counts["audit_kpi_master"] = n

    db.commit()
    logger.info("process_group masters seeded: %s", counts)
    return counts


def list_process_groups(db: Session) -> List[Dict[str, Any]]:
    ensure_process_group_tables(db)
    if not _table_exists(db, "process_group_master"):
        return []
    rows = db.execute(
        text(
            "SELECT process_group_id, process_group_name, hls_scope_desc "
            "FROM process_group_master ORDER BY process_group_id"
        )
    ).mappings().all()
    return [dict(r) for r in rows]


def list_process_group_tree(
    db: Session, standard_code: Optional[str] = None
) -> List[Dict[str, Any]]:
    """Process groups with HLS codes + optional standard-specific clause maps."""
    groups = list_process_groups(db)
    if not groups:
        return []

    hls_by_pg: Dict[str, List[Dict[str, Any]]] = {}
    if _table_exists(db, "process_group_hls_map") and _table_exists(db, "hls_master"):
        rows = db.execute(
            text(
                "SELECT m.process_group_id, m.hls_code, h.checkpoints_summary "
                "FROM process_group_hls_map m "
                "JOIN hls_master h ON h.hls_code = m.hls_code "
                "ORDER BY m.process_group_id, m.hls_code"
            )
        ).mappings().all()
        for r in rows:
            hls_by_pg.setdefault(r["process_group_id"], []).append(
                {
                    "hls_code": r["hls_code"],
                    "checkpoints_summary": r["checkpoints_summary"],
                }
            )

    clauses_by_pg: Dict[str, List[Dict[str, Any]]] = {}
    if standard_code and _table_exists(db, "standard_process_clause_map"):
        rows = db.execute(
            text(
                "SELECT process_group_id, actual_clause_no, clause_topic, guide_note "
                "FROM standard_process_clause_map "
                "WHERE standard_code = :std "
                "ORDER BY process_group_id, actual_clause_no"
            ),
            {"std": standard_code},
        ).mappings().all()
        for r in rows:
            clauses_by_pg.setdefault(r["process_group_id"], []).append(
                {
                    "actual_clause_no": r["actual_clause_no"],
                    "clause_topic": r["clause_topic"],
                    "guide_note": r["guide_note"],
                }
            )

    out = []
    for g in groups:
        pg = g["process_group_id"]
        item = {
            **g,
            "hls_codes": hls_by_pg.get(pg, []),
        }
        if standard_code:
            item["standard_clauses"] = clauses_by_pg.get(pg, [])
        out.append(item)
    return out


def list_audit_kpis(
    db: Session,
    *,
    hls_code: Optional[str] = None,
    standard_code: Optional[str] = None,
) -> List[Dict[str, Any]]:
    ensure_process_group_tables(db)
    if not _table_exists(db, "audit_kpi_master"):
        return []
    sql = (
        "SELECT kpi_id, hls_code, standard_code, kpi_name, kpi_type, formula, unit "
        "FROM audit_kpi_master WHERE 1=1"
    )
    params: Dict[str, Any] = {}
    if hls_code:
        sql += " AND hls_code = :hls"
        params["hls"] = hls_code
    if standard_code:
        sql += " AND (standard_code = :std OR standard_code = 'COMMON')"
        params["std"] = standard_code
    sql += " ORDER BY hls_code, kpi_id"
    return [dict(r) for r in db.execute(text(sql), params).mappings().all()]


def list_standard_masters_pg(db: Session) -> List[Dict[str, Any]]:
    """Process-schema standard_master (not platform standard_masters)."""
    ensure_process_group_tables(db)
    if not _table_exists(db, "standard_master"):
        return []
    rows = db.execute(
        text(
            "SELECT standard_code, standard_name, hls_adopted, native_structure_note "
            "FROM standard_master ORDER BY standard_code"
        )
    ).mappings().all()
    return [dict(r) for r in rows]


def count_clauses_for_standard_pg(db: Session, standard_code: str) -> int:
    return len(list_clauses_for_standard_pg(db, standard_code))


def list_clauses_for_standard_pg(
    db: Session,
    standard_code: str,
    *,
    standard_key: Optional[str] = None,
) -> List[Dict[str, Any]]:
    """Build audit-note clause nav from process-group masters for one standard.

    Primary: process_group_hls_map × hls_master × standard_clause_map
    Plus: standard_process_clause_map (standard-specific operational clauses)
    Titles: clause_topic → HLS title fallback (never iso_clauses_master)
    KPIs: audit_kpi_master by hls_code / range (empty values allowed at save time)
    """
    ensure_process_group_tables(db)
    pg_code = to_process_standard_code(standard_code) or standard_code
    if not pg_code or not _table_exists(db, "process_group_master"):
        return []

    groups = list_process_groups(db)
    if not groups:
        return []

    # HLS map rows + full hls_master checkpoint lookup (reference guide only)
    hls_rows = []
    hls_cp_map: Dict[str, Optional[str]] = {}
    if _table_exists(db, "hls_master"):
        for hr in db.execute(
            text("SELECT hls_code, checkpoints_summary FROM hls_master")
        ).mappings().all():
            hls_cp_map[str(hr["hls_code"])] = hr.get("checkpoints_summary")
    if _table_exists(db, "process_group_hls_map") and _table_exists(db, "hls_master"):
        hls_rows = db.execute(
            text(
                "SELECT m.process_group_id, m.hls_code, h.checkpoints_summary "
                "FROM process_group_hls_map m "
                "JOIN hls_master h ON h.hls_code = m.hls_code "
                "ORDER BY m.process_group_id, m.hls_code"
            )
        ).mappings().all()

    clause_map: Dict[str, Dict[str, Any]] = {}
    if _table_exists(db, "standard_clause_map"):
        for r in db.execute(
            text(
                "SELECT hls_code, actual_clause_no, status, "
                "integrated_into_hls_code, guide_note "
                "FROM standard_clause_map WHERE standard_code = :std"
            ),
            {"std": pg_code},
        ).mappings().all():
            clause_map[r["hls_code"]] = dict(r)

    proc_clauses = []
    if _table_exists(db, "standard_process_clause_map"):
        proc_clauses = [
            dict(r)
            for r in db.execute(
                text(
                    "SELECT process_group_id, actual_clause_no, clause_topic, guide_note "
                    "FROM standard_process_clause_map WHERE standard_code = :std "
                    "ORDER BY process_group_id, actual_clause_no"
                ),
                {"std": pg_code},
            ).mappings().all()
        ]

    kpis_all = list_audit_kpis(db, standard_code=pg_code)

    # Titles come ONLY from process-group masters (clause_topic / HLS titles).
    # Do NOT use iso_clauses_master — it injects EMS 환경심층 (6.1c TCFD 등) into other standards.
    pg_name = {g["process_group_id"]: g["process_group_name"] for g in groups}
    seen: set = set()
    out: List[Dict[str, Any]] = []
    sort_i = 0

    def _kpis_for(hls_code: str, clause_no: str) -> List[Dict[str, str]]:
        items = []
        for k in kpis_all:
            if _hls_matches_clause(k["hls_code"], hls_code, clause_no):
                kid = k["kpi_id"]
                kn = k["kpi_name"]
                items.append(
                    {
                        "kpi_id": kid,
                        "kpi_name": kn,
                        "key": kid,
                        "label": kn,
                    }
                )
        return items

    def _title_for(clause_no: str, hls_code: str, topic: Optional[str]) -> str:
        raw = (topic or "").strip()
        if not raw:
            raw = _HLS_TITLE_KO.get(hls_code) or _HLS_TITLE_KO.get(clause_no.split("/")[0].strip()) or ""
        return _clean_clause_title(clause_no, raw)

    def _cps_for(hls_code: str, summary: Optional[str], guide: Optional[str]) -> List[Dict[str, str]]:
        """Checkpoints from hls_master.checkpoints_summary (+ optional guide_note)."""
        raw = summary if summary is not None else hls_cp_map.get(hls_code)
        if raw is None and hls_code:
            # parent digit.digit fallback e.g. 8.2 → 8.1 range keys already in map
            head = hls_code.split("/")[0].strip()
            raw = hls_cp_map.get(head)
        cps = _split_checkpoints(raw)
        if guide and str(guide).strip():
            cps = [{"title": "표준별 가이드", "hint": str(guide).strip()}] + cps
        return _filter_checkpoints_for_standard(cps, pg_code)

    # A) HLS-driven rows by process group
    for r in hls_rows:
        hls = r["hls_code"]
        cmap = clause_map.get(hls) or {}
        status = (cmap.get("status") or "DIRECT").upper()
        if status == "INTEGRATED":
            # Not a standalone clause for this standard
            continue
        clause_no = (cmap.get("actual_clause_no") or hls or "").strip()
        if not clause_no or clause_no in seen:
            continue
        seen.add(clause_no)
        sort_i += 1
        guide = cmap.get("guide_note")
        cps = _cps_for(hls, r.get("checkpoints_summary"), guide)
        topic = _title_for(clause_no, hls, None)
        pgid = r["process_group_id"]
        out.append(
            {
                "id": sort_i,
                "standard_key": standard_key or pg_code,
                "standard_code": pg_code,
                "family_code": None,
                "clause_no": clause_no,
                "clause_topic": topic,
                "clause_title": topic,
                "question": (guide or ""),
                "default_kpis": _kpis_for(hls, clause_no),
                "checkpoints": cps,
                "process_group_name": pg_name.get(pgid) or pgid,
                "group_name": pg_name.get(pgid) or pgid,
                "process_group_id": pgid,
                "hls_code": hls,
                "sort_order": sort_i,
                "source": "process_group",
            }
        )

    # B) Standard-specific process clauses (8.2/8.3/…) — also pull hls_master checkpoints
    for r in proc_clauses:
        clause_no = (r.get("actual_clause_no") or "").strip()
        if not clause_no or clause_no in seen:
            # If already present, enrich title/topic when empty
            if clause_no in seen and r.get("clause_topic"):
                for item in out:
                    if item["clause_no"] == clause_no and not item.get("clause_topic"):
                        item["clause_topic"] = r["clause_topic"]
                        item["clause_title"] = r["clause_topic"]
            continue
        seen.add(clause_no)
        sort_i += 1
        hls = clause_no.split("/")[0].strip()
        guide = r.get("guide_note")
        cps = _cps_for(hls, None, guide)
        topic = _title_for(clause_no, hls, r.get("clause_topic"))
        pgid = r["process_group_id"]
        out.append(
            {
                "id": sort_i,
                "standard_key": standard_key or pg_code,
                "standard_code": pg_code,
                "family_code": None,
                "clause_no": clause_no,
                "clause_topic": topic,
                "clause_title": topic,
                "question": (guide or ""),
                "default_kpis": _kpis_for(hls, clause_no),
                "checkpoints": cps,
                "process_group_name": pg_name.get(pgid) or pgid,
                "group_name": pg_name.get(pgid) or pgid,
                "process_group_id": pgid,
                "hls_code": hls,
                "sort_order": sort_i,
                "source": "process_group",
            }
        )

    out.sort(
        key=lambda x: (
            x.get("process_group_id") or "",
            _clause_sort_key(x.get("clause_no") or ""),
        )
    )
    for i, item in enumerate(out, start=1):
        item["id"] = i
        item["sort_order"] = i

    # Dual KPI panels: ISO (excel chapter + HLS audit_kpi) + ESG master
    try:
        from app.services.iso_audit_kpis import enrich_clause_dual_kpis

        for item in out:
            enrich_clause_dual_kpis(db, item)
    except Exception:
        logger.exception("enrich dual KPIs soft-fail for %s", pg_code)
        for item in out:
            item.setdefault("iso_audit_kpis", list(item.get("default_kpis") or []))
            item.setdefault("esg_kpis", [])

    return out
